import sys
import os
import csv
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def validate_file(filename):
    if not filename.endswith(".csv"):
        print("ERROR: File must be CSV")
        sys.exit(1)

    if not os.path.exists(filename):
        print("ERROR: File not found")
        sys.exit(1)


def read_dataset(filename):
    data = []

    with open(filename, mode="r", encoding="utf-8") as file:
        csv_reader = csv.DictReader(file)

        for row in csv_reader:
            data.append(row)

    return data


def get_numeric_value(row, column_name):
    try:
        return float(row[column_name])
    except KeyError:
        return None
    except ValueError:
        return None


def calculate_average(dataset, column_name):
    total = 0
    count = 0

    for row in dataset:
        value = get_numeric_value(row, column_name)

        if value is not None:
            total += value
            count += 1

    if count == 0:
        return 0

    return total / count


def calculate_sales_by_column(dataset, group_by_column):
    sales_by_column = {}

    for row in dataset:
        if group_by_column not in row:
            continue

        group_value = row[group_by_column]
        global_sales = get_numeric_value(row, "Global_Sales")

        if global_sales is None:
            continue

        if group_value not in sales_by_column:
            sales_by_column[group_value] = 0

        sales_by_column[group_value] += global_sales

    return sales_by_column


def calculate_summary(dataset):
    total_games = len(dataset)
    total_global_sales = 0
    best_selling_game = ""
    best_selling_game_sales = 0

    for row in dataset:
        global_sales = get_numeric_value(row, "Global_Sales")

        if global_sales is None:
            continue

        total_global_sales += global_sales

        if global_sales > best_selling_game_sales:
            best_selling_game_sales = global_sales
            best_selling_game = row["Name"]

    summary = {
        "total_games": total_games,
        "total_global_sales": total_global_sales,
        "best_selling_game": best_selling_game,
        "best_selling_game_sales": best_selling_game_sales
    }

    return summary


def display_summary(dataset):
    summary = calculate_summary(dataset)

    print("\n--- Summary ---")
    print(f"Total number of games: {summary['total_games']}")
    print(f"Total global sales: {summary['total_global_sales']:.2f} million")
    print(
        f"Best-selling game: {summary['best_selling_game']} "
        f"({summary['best_selling_game_sales']:.2f} million)"
    )


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")


def save_excel_report(dataset, stat_column, group_by_column, output_file):
    summary = calculate_summary(dataset)
    average_value = calculate_average(dataset, stat_column)
    grouped_sales = calculate_sales_by_column(dataset, group_by_column)

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total number of games", summary["total_games"]])
    summary_sheet.append([
        "Total global sales",
        round(summary["total_global_sales"], 2)
    ])
    summary_sheet.append(["Best-selling game", summary["best_selling_game"]])
    summary_sheet.append([
        "Best-selling game sales",
        round(summary["best_selling_game_sales"], 2)
    ])

    style_header(summary_sheet[1])

    statistics_sheet = workbook.create_sheet("Statistics")
    statistics_sheet.append(["Statistic", "Column", "Value"])
    statistics_sheet.append([
        "Average",
        stat_column,
        round(average_value, 2)
    ])

    style_header(statistics_sheet[1])

    aggregation_sheet = workbook.create_sheet("Aggregation")
    aggregation_sheet.append([group_by_column, "Total Global Sales"])

    sorted_grouped_sales = sorted(
        grouped_sales.items(),
        key=lambda item: item[1],
        reverse=True
    )

    for group, sales in sorted_grouped_sales:
        aggregation_sheet.append([group, round(sales, 2)])

    style_header(aggregation_sheet[1])

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                cell_value = str(cell.value)

                if len(cell_value) > max_length:
                    max_length = len(cell_value)

            sheet.column_dimensions[column_letter].width = max_length + 4

    workbook.save(output_file)


def display_help():
    print("Usage:")
    print("  python app8.py dataset.csv")
    print("  python app8.py dataset.csv -o report.xlsx")
    print()
    print("Purpose:")
    print("  This application reads a CSV dataset with video game sales data,")
    print("  calculates summary, statistical and aggregation results,")
    print("  and optionally saves them to an Excel report.")
    print()
    print("Arguments:")
    print("  dataset.csv       Required CSV dataset file")
    print()
    print("Options:")
    print("  -o report.xlsx    Save all results to an Excel file")
    print("  -h                Display this help message")
    print()
    print("Environment variables:")
    print("  STAT_COLUMN       Numeric column used for average calculation")
    print("  GROUP_BY_COLUMN   Column used for aggregation")


def parse_arguments():
    if len(sys.argv) == 2 and sys.argv[1] == "-h":
        display_help()
        sys.exit(0)

    if len(sys.argv) < 2:
        print("Usage: python app8.py dataset.csv [-o report.xlsx]")
        print("Use python app8.py -h for help.")
        sys.exit(1)

    filename = sys.argv[1]
    output_file = None

    if "-o" in sys.argv:
        output_index = sys.argv.index("-o")

        if output_index + 1 >= len(sys.argv):
            print("ERROR: Missing output file name after -o")
            sys.exit(1)

        output_file = sys.argv[output_index + 1]

    return filename, output_file


def main():
    load_dotenv()

    filename, output_file = parse_arguments()

    validate_file(filename)

    stat_column = os.getenv("STAT_COLUMN", "Global_Sales")
    group_by_column = os.getenv("GROUP_BY_COLUMN", "Genre")

    dataset = read_dataset(filename)

    print("Dataset is valid!")
    print(f"Loaded {len(dataset)} rows")

    if output_file is None:
        display_summary(dataset)
    else:
        save_excel_report(dataset, stat_column, group_by_column, output_file)
        print(f"Excel report saved to {output_file}")


if __name__ == "__main__":
    main()
